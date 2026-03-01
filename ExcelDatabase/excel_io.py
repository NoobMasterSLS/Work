import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
from datetime import datetime, date
from config import DATA_FOLDER, TEMPLATE_FILE
from helpers import sanitize_filename

# константы для столбцов сотрудников (B..I)
EMPLOYEE_COLS = list(range(2, 10))  # индексы столбцов: B=2, C=3, ..., I=9
EMPLOYEE_HEADERS = ['Фамилия', 'Имя', 'Отчество', 'Дата рождения',
                    'Дата трудоустройства', 'Телефон', 'Должность', 'Оклад']

def auto_fit_columns(worksheet):
    """автоматически подбирает ширину столбцов по содержимому"""
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[col_letter].width = min(adjusted_width, 50)

def read_excel_and_save_to_json(excel_path):
    """
    читает Excel-файл, извлекает данные и сохраняет в JSON
    возвращает название компании
    """
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # поиск строки с "Информация о сотрудниках"
    info_row = None
    for row in range(1, sheet.max_row + 1):
        cell_val = sheet.cell(row, 1).value
        if cell_val and "Информация о сотрудниках" in str(cell_val):
            info_row = row
            break
    if info_row is None:
        raise ValueError("В файле не найдена строка с 'Информация о сотрудниках'")

    # общая инфа
    company_info = {}
    for row in range(2, info_row):
        field_name = sheet.cell(row, 1).value
        field_value = sheet.cell(row, 2).value  # столбец B
        if field_name:
            if isinstance(field_value, (datetime, date)):
                field_value = field_value.strftime('%d.%m.%Y')
            elif field_value is None:
                field_value = ''
            else:
                field_value = str(field_value)
            company_info[field_name] = field_value

    company_name = company_info.get('Название')
    if not company_name:
        raise ValueError("Не удалось определить название компании (поле 'Название' не найдено или пусто)")

    # --- чтение сотрудников ---
    employees = []
    start_row = info_row + 1

    # определяем, есть ли строка с заголовками
    header_row_candidate = sheet.cell(start_row, 2).value  # B
    if header_row_candidate and header_row_candidate.strip() in EMPLOYEE_HEADERS:
        start_data_row = start_row + 1
    else:
        start_data_row = start_row

    # проходим по всем строкам от start_data_row до конца листа
    max_row = sheet.max_row
    for row in range(start_data_row, max_row + 1):
        # проверяем, есть ли хоть одно непустое значение в столбцах B..I
        has_data = False
        emp_data = {}
        for col_idx in EMPLOYEE_COLS:
            cell = sheet.cell(row, col_idx)
            val = cell.value
            if val is not None and str(val).strip() != '':
                has_data = True
            if isinstance(val, (datetime, date)):
                val = val.strftime('%d.%m.%Y')
            elif val is None:
                val = ''
            else:
                val = str(val)
            emp_data[EMPLOYEE_HEADERS[col_idx - 2]] = val

        if has_data:
            # номер сотрудника (столбец A)
            emp_num = sheet.cell(row, 1).value
            if emp_num is None:
                emp_num = ''
            else:
                emp_num = str(emp_num)

            employee_record = {
                'employee_number': emp_num,
                'last_name': emp_data.get('Фамилия', ''),
                'first_name': emp_data.get('Имя', ''),
                'middle_name': emp_data.get('Отчество', ''),
                'birth_date': emp_data.get('Дата рождения', ''),
                'hire_date': emp_data.get('Дата трудоустройства', ''),
                'phone': emp_data.get('Телефон', ''),
                'position': emp_data.get('Должность', ''),
                'salary': emp_data.get('Оклад', '')
            }
            employees.append(employee_record)

    # сохраняем в JSON
    data = {
        'company_info': company_info,
        'employees': employees
    }
    safe_name = sanitize_filename(company_name)
    json_path = os.path.join(DATA_FOLDER, f'{safe_name}.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

    print(f"✅ Данные компании '{company_name}' сохранены в файл: {json_path}")
    return company_name

def load_from_json_and_create_excel(company_name, output_excel_path=None):
    """
    загружает данные компании из JSON и создаёт Excel-файл на основе шаблона
    """
    safe_name = sanitize_filename(company_name)
    json_path = os.path.join(DATA_FOLDER, f'{safe_name}.json')
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"Компания '{company_name}' не найдена в базе.")

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    company_info = data['company_info']
    employees = data['employees']

    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Файл шаблона '{TEMPLATE_FILE}' не найден в текущей папке.")

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    sheet = wb.active

    # заполнение общей информации
    for row in range(1, sheet.max_row + 1):
        field_name = sheet.cell(row, 1).value
        if field_name and field_name in company_info:
            sheet.cell(row, 2).value = company_info[field_name]

    # поиск строки "Информация о сотрудниках"
    info_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value and "Информация о сотрудниках" in str(sheet.cell(row, 1).value):
            info_row = row
            break
    if info_row is None:
        raise ValueError("В шаблоне не найдена строка с 'Информация о сотрудниках'")

    # определяем начало данных
    start_row = info_row + 1
    header_row = None
    first_cell_val = sheet.cell(start_row, 2).value
    if first_cell_val and first_cell_val.strip() in EMPLOYEE_HEADERS:
        header_row = start_row
        start_data_row = start_row + 1
    else:
        start_data_row = start_row

    # очищаем старые данные
    max_row = sheet.max_row
    for r in range(start_data_row, max_row + 1):
        for c in EMPLOYEE_COLS:
            sheet.cell(r, c).value = None
        sheet.cell(r, 1).value = None

    # записываем сотрудников
    for i, emp in enumerate(employees):
        current_row = start_data_row + i
        sheet.cell(current_row, 1).value = i + 1
        sheet.cell(current_row, 2).value = emp.get('last_name', '')
        sheet.cell(current_row, 3).value = emp.get('first_name', '')
        sheet.cell(current_row, 4).value = emp.get('middle_name', '')
        sheet.cell(current_row, 5).value = emp.get('birth_date', '')
        sheet.cell(current_row, 6).value = emp.get('hire_date', '')
        sheet.cell(current_row, 7).value = emp.get('phone', '')
        sheet.cell(current_row, 8).value = emp.get('position', '')
        sheet.cell(current_row, 9).value = emp.get('salary', '')

    auto_fit_columns(sheet)

    if output_excel_path is None:
        output_excel_path = os.path.join(DATA_FOLDER, f'{safe_name}_output.xlsx')
    wb.save(output_excel_path)
    print(f"✅ Excel-файл создан: {output_excel_path}")
    return output_excel_path

def read_json_company(company_name):
    """
    читает данные компании из JSON-файла и возвращает словарь с company_info и employees.
    """
    safe_name = sanitize_filename(company_name)
    json_path = os.path.join(DATA_FOLDER, f'{safe_name}.json')
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"Компания '{company_name}' не найдена в базе.")

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data