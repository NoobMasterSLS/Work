import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, date
from constants import TEMPLATE_FILE
from companies.models import Company, Employee
from django.db.utils import OperationalError

# константы для столбцов сотрудников (B..I)
EMPLOYEE_COLS = list(range(2, 10))  # индексы столбцов: B=2, C=3, ..., I=9
EMPLOYEE_HEADERS = ['Фамилия', 'Имя', 'Отчество', 'Дата рождения',
                    'Дата трудоустройства', 'Телефон', 'Должность', 'Оклад']

def auto_fit_columns(worksheet):
    """автоматически подбирает ширину столбцов по содержимому"""
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[col_letter].width = min(adjusted_width, 50)

def read_excel_and_save_to_db(excel_path):
    """
    читает Excel-файл, извлекает данные и сохраняет в БД через Django ORM
    возвращает название компании или None в случае ошибки БД
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # поиск строки с "Информация о сотрудниках"
        info_row = None
        for row in range(1, sheet.max_row + 1):
            cell_val = sheet.cell(row, 1).value
            if cell_val and "Информация о сотрудниках" in str(cell_val):
                info_row = row
                break
        if info_row is None:
            raise ValueError("В файле не найдена строка с 'Информация о сотрудниках'")

        # общая инфа
        company_info = {}
        for row in range(2, info_row):
            field_name = sheet.cell(row, 1).value
            field_value = sheet.cell(row, 2).value
            if field_name:
                if isinstance(field_value, (datetime, date)):
                    field_value = field_value.strftime('%d.%m.%Y')
                elif field_value is None:
                    field_value = ''
                else:
                    field_value = str(field_value)
                company_info[field_name] = field_value

        company_name = company_info.get('Название')
        if not company_name:
            raise ValueError("Не удалось определить название компании (поле 'Название' не найдено или пусто)")

        # Создаём или обновляем компанию
        company, created = Company.objects.update_or_create(
            name=company_name,
            defaults={'extra_data': company_info}
        )

        # --- чтение сотрудников ---
        employees = []
        start_row = info_row + 1

        # определяем, есть ли строка с заголовками
        header_row_candidate = sheet.cell(start_row, 2).value
        if header_row_candidate and header_row_candidate.strip() in EMPLOYEE_HEADERS:
            start_data_row = start_row + 1
        else:
            start_data_row = start_row

        # проходим по всем строкам от start_data_row до конца листа
        max_row = sheet.max_row
        for row in range(start_data_row, max_row + 1):
            # проверяем, есть ли хоть одно непустое значение в столбцах B..I
            has_data = False
            emp_data = {}
            for col_idx in EMPLOYEE_COLS:
                cell = sheet.cell(row, col_idx)
                val = cell.value
                if val is not None and str(val).strip() != '':
                    has_data = True
                if isinstance(val, (datetime, date)):
                    val = val.strftime('%d.%m.%Y')
                elif val is None:
                    val = ''
                else:
                    val = str(val)
                emp_data[EMPLOYEE_HEADERS[col_idx - 2]] = val

            if has_data:
                employee_record = {
                    'last_name': emp_data.get('Фамилия', ''),
                    'first_name': emp_data.get('Имя', ''),
                    'middle_name': emp_data.get('Отчество', ''),
                    'birth_date': emp_data.get('Дата рождения', ''),
                    'hire_date': emp_data.get('Дата трудоустройства', ''),
                    'phone': emp_data.get('Телефон', ''),
                    'position': emp_data.get('Должность', ''),
                    'salary': emp_data.get('Оклад', '')
                }
                employees.append(employee_record)

        # Удаляем старых сотрудников компании и создаём новых
        company.employees.all().delete()
        for emp_data in employees:
            Employee.objects.create(company=company, **emp_data)

        print(f"✅ Данные компании '{company_name}' сохранены в БД.")
        return company_name

    except OperationalError as e:
        if 'no such table' in str(e):
            print("\n❌ База данных не инициализирована. Выполните миграции:")
            print("   python manage.py migrate\n")
            return None
        else:
            raise

def load_from_db_and_create_excel(company_name, output_excel_path=None):
    """
    загружает данные компании из БД и создаёт Excel-файл на основе шаблона
    """
    try:
        company = Company.objects.get(name=company_name)
    except OperationalError as e:
        if 'no such table' in str(e):
            print("\n❌ База данных не инициализирована. Выполните миграции:")
            print("   python manage.py migrate\n")
            return None
        else:
            raise
    except Company.DoesNotExist:
        raise FileNotFoundError(f"Компания '{company_name}' не найдена в БД.")

    company_info = company.extra_data
    employees = company.employees.all()

    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Файл шаблона '{TEMPLATE_FILE}' не найден в текущей папке.")

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    sheet = wb.active

    # заполнение общей информации
    for row in range(1, sheet.max_row + 1):
        field_name = sheet.cell(row, 1).value
        if field_name and field_name in company_info:
            sheet.cell(row, 2).value = company_info[field_name]

    # поиск строки "Информация о сотрудниках"
    info_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value and "Информация о сотрудниках" in str(sheet.cell(row, 1).value):
            info_row = row
            break
    if info_row is None:
        raise ValueError("В шаблоне не найдена строка с 'Информация о сотрудниках'")

    # определяем начало данных
    start_row = info_row + 1
    header_row = None
    first_cell_val = sheet.cell(start_row, 2).value
    if first_cell_val and first_cell_val.strip() in EMPLOYEE_HEADERS:
        header_row = start_row
        start_data_row = start_row + 1
    else:
        start_data_row = start_row

    # очищаем старые данные
    max_row = sheet.max_row
    for r in range(start_data_row, max_row + 1):
        for c in EMPLOYEE_COLS:
            sheet.cell(r, c).value = None
        sheet.cell(r, 1).value = None

    # записываем сотрудников
    for i, emp in enumerate(employees):
        current_row = start_data_row + i
        sheet.cell(current_row, 1).value = i + 1
        sheet.cell(current_row, 2).value = emp.last_name
        sheet.cell(current_row, 3).value = emp.first_name
        sheet.cell(current_row, 4).value = emp.middle_name
        sheet.cell(current_row, 5).value = emp.birth_date
        sheet.cell(current_row, 6).value = emp.hire_date
        sheet.cell(current_row, 7).value = emp.phone
        sheet.cell(current_row, 8).value = emp.position
        sheet.cell(current_row, 9).value = emp.salary

    auto_fit_columns(sheet)

    if output_excel_path is None:
        # генерируем имя по умолчанию
        safe_name = company_name.replace('/', '_').replace('\\', '_')
        output_excel_path = os.path.join(os.getcwd(), f'{safe_name}_output.xlsx')
    wb.save(output_excel_path)
    print(f"✅ Excel-файл создан: {output_excel_path}")
    return output_excel_path

def display_company_info_from_db(company_name):
    """
    выводит информацию о компании и сотрудниках в консоль
    """
    try:
        company = Company.objects.get(name=company_name)
    except OperationalError as e:
        if 'no such table' in str(e):
            print("\n❌ База данных не инициализирована. Выполните миграции:")
            print("   python manage.py migrate\n")
            return
        else:
            raise
    except Company.DoesNotExist:
        print(f"❌ Компания '{company_name}' не найдена в БД.")
        return

    company_info = company.extra_data
    employees = company.employees.all()

    print("\n" + "="*60)
    print(f"ИНФОРМАЦИЯ О КОМПАНИИ: {company_name}")
    print("="*60)
    print("\n--- Общая информация ---")
    for key, value in company_info.items():
        print(f"{key}: {value}")

    print(f"\n--- Сотрудники ({employees.count()}) ---")
    if not employees:
        print("Нет данных о сотрудниках.")
    else:
        headers = ['№', 'Фамилия', 'Имя', 'Отчество', 'Дата рожд.', 'Дата приёма', 'Телефон', 'Должность', 'Оклад']
        col_widths = [5, 15, 10, 15, 12, 12, 12, 15, 10]
        header_line = ' '.join([f"{h:<{w}}" for h, w in zip(headers, col_widths)])
        print(header_line)
        print('-' * len(header_line))

        for i, emp in enumerate(employees, 1):
            row = [
                str(i),
                emp.last_name[:15],
                emp.first_name[:10],
                emp.middle_name[:15],
                emp.birth_date[:12],
                emp.hire_date[:12],
                emp.phone[:12],
                emp.position[:15],
                emp.salary[:10]
            ]
            row_line = ' '.join([f"{str(cell):<{w}}" for cell, w in zip(row, col_widths)])
            print(row_line)
    print("="*60)